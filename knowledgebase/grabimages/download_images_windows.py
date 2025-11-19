#!/usr/bin/env python3
"""
Download images from ManageEngine web interface and embed them into Excel
Windows-compatible version
"""
import csv
import re
import os
import sys
import time
import requests
from requests.auth import HTTPBasicAuth
import urllib3
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from io import BytesIO
from PIL import Image
import platform

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Set CSV field size limit (Windows-compatible)
try:
    if platform.system() == 'Windows':
        csv.field_size_limit(500000)  # Set to a reasonable limit for Windows
    else:
        csv.field_size_limit(sys.maxsize)
except:
    csv.field_size_limit(500000)  # Fallback to safe value

# ManageEngine credentials and URL
HELPDESK_URL = "https://helpdesk.appsforte.com"
# You'll need to provide credentials
print("=" * 70)
print("ManageEngine ServiceDesk Image Downloader - Windows Edition")
print("=" * 70)
username = input("Enter ManageEngine username: ").strip()
password = input("Enter ManageEngine password: ").strip()
print()

def extract_embedded_images_from_csv():
    """Extract all embedded image references from solutions CSV"""
    print("[1/4] Extracting embedded image references...")

    image_refs = {}  # solution_id -> list of image_ids

    # Determine the path to solutions.csv
    script_dir = os.path.dirname(os.path.abspath(__file__))
    solutions_csv = os.path.join(script_dir, 'solutions.csv')

    # Try current directory if not found
    if not os.path.exists(solutions_csv):
        solutions_csv = 'solutions.csv'

    if not os.path.exists(solutions_csv):
        print(f"      ERROR: Cannot find solutions.csv")
        print(f"      Looked in: {script_dir}")
        return {}

    with open(solutions_csv, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find all solutions with their IDs and embedded images
    lines = content.split('\n')
    for line in lines:
        if not line.strip():
            continue

        # Extract solution ID (first field)
        match = re.match(r'^(\d+),', line)
        if not match:
            continue

        sol_id = match.group(1)

        # Find all image references in this line
        img_matches = re.findall(r'/api/v3/solutions/\d+/images/(\d+)', line)

        if img_matches:
            image_refs[sol_id] = list(set(img_matches))  # Remove duplicates

    total_images = sum(len(imgs) for imgs in image_refs.values())
    print(f"      Found {total_images} images across {len(image_refs)} solutions")
    print()

    return image_refs

def download_image(session, solution_id, image_id):
    """Download a single image from ManageEngine"""
    urls_to_try = [
        f"{HELPDESK_URL}/api/v3/solutions/{solution_id}/images/{image_id}",
        f"{HELPDESK_URL}/solutions/{solution_id}/images/{image_id}",
    ]

    for url in urls_to_try:
        try:
            response = session.get(url, verify=False, timeout=30)

            if response.status_code == 200 and len(response.content) > 100:
                return response.content
        except Exception as e:
            continue

    return None

def save_images_locally(image_refs, session):
    """Download all images and save them locally"""
    # Create images directory in script location or current directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    image_dir = os.path.join(script_dir, "downloaded_images")

    if not os.path.exists(image_dir):
        os.makedirs(image_dir)

    print(f"[2/4] Downloading images from {HELPDESK_URL}...")
    print(f"      Saving to: {image_dir}")
    print(f"      This may take several minutes...")
    print()

    downloaded_images = {}  # (sol_id, img_id) -> local_path
    total_images = sum(len(imgs) for imgs in image_refs.values())
    current = 0
    successful = 0
    failed = 0

    for sol_id, img_ids in sorted(image_refs.items()):
        for img_id in img_ids:
            current += 1

            # Show progress
            if current % 10 == 0 or current == 1:
                print(f"      [{current}/{total_images}] Downloading... ({successful} OK, {failed} failed)")

            img_data = download_image(session, sol_id, img_id)

            if img_data:
                # Save to file
                filename = f"solution_{sol_id}_image_{img_id}.png"
                filepath = os.path.join(image_dir, filename)

                try:
                    with open(filepath, 'wb') as f:
                        f.write(img_data)

                    downloaded_images[(sol_id, img_id)] = filepath
                    successful += 1
                except:
                    failed += 1
            else:
                failed += 1

            # Be nice to the server
            time.sleep(0.3)

    print(f"      [{current}/{total_images}] Complete!")
    print()
    print(f"      Downloaded: {successful}/{total_images} images")
    print(f"      Failed: {failed}/{total_images}")
    print()

    return downloaded_images

def create_excel_with_images(downloaded_images, image_refs):
    """Create Excel file with embedded images"""
    print("[3/4] Creating Excel spreadsheet with embedded images...")

    # Load solutions
    print("      Loading solution data...")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    solutions_csv = os.path.join(script_dir, 'solutions.csv')
    if not os.path.exists(solutions_csv):
        solutions_csv = 'solutions.csv'

    solutions = {}
    with open(solutions_csv, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            solutions[row['solutionid']] = row

    # Load attachments
    attachments_csv = os.path.join(script_dir, 'solution_attachments.csv')
    if not os.path.exists(attachments_csv):
        attachments_csv = 'solution_attachments.csv'

    attachments = defaultdict(list)
    try:
        with open(attachments_csv, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                attachments[row['solutionid']].append({
                    'filename': row['attachmentname'],
                })
    except:
        pass

    # Create workbook
    print("      Creating Excel workbook...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KB with Images"

    # Headers
    headers = ['ID', 'Title', 'Description', 'Images', 'Attachments']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25

    # Add data rows
    print("      Adding solutions and embedding images...")
    row_num = 2
    images_embedded = 0

    for sol_id in sorted(solutions.keys(), key=lambda x: int(x)):
        sol = solutions[sol_id]

        # Clean description (remove HTML tags)
        desc_text = re.sub(r'<[^>]+>', '', sol['description'])
        desc_text = re.sub(r'\s+', ' ', desc_text).strip()
        if len(desc_text) > 400:
            desc_text = desc_text[:400] + "..."

        # Basic info
        ws.cell(row=row_num, column=1, value=int(sol_id))
        ws.cell(row=row_num, column=2, value=sol['title'])
        ws.cell(row=row_num, column=3, value=desc_text)

        # Add images in column D
        if sol_id in image_refs:
            img_ids = image_refs[sol_id]
            ws.cell(row=row_num, column=4, value=f"{len(img_ids)} images")

            # Try to embed first image
            first_img_id = img_ids[0]
            img_key = (sol_id, first_img_id)

            if img_key in downloaded_images:
                try:
                    img_path = downloaded_images[img_key]

                    # Resize image if needed
                    pil_img = Image.open(img_path)
                    max_width = 180
                    max_height = 130

                    if pil_img.width > max_width or pil_img.height > max_height:
                        pil_img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                        resized_path = img_path.replace('.png', '_resized.png')
                        pil_img.save(resized_path)
                        img_path = resized_path

                    # Add to Excel
                    excel_img = ExcelImage(img_path)
                    excel_img.anchor = f'D{row_num}'
                    ws.add_image(excel_img)

                    # Adjust row height
                    ws.row_dimensions[row_num].height = 100
                    images_embedded += 1
                except:
                    pass
        else:
            ws.cell(row=row_num, column=4, value="No images")

        # Attachments
        sol_attachments = attachments.get(sol_id, [])
        if sol_attachments:
            att_text = '\n'.join([a['filename'] for a in sol_attachments[:3]])
            if len(sol_attachments) > 3:
                att_text += f"\n...+{len(sol_attachments)-3}"
            ws.cell(row=row_num, column=5, value=att_text)
        else:
            ws.cell(row=row_num, column=5, value="-")

        row_num += 1

        if row_num % 100 == 0:
            print(f"      Processed {row_num-1} solutions...")

    # Freeze header
    ws.freeze_panes = 'A2'

    # Save in same directory as script or current directory
    output_file = os.path.join(script_dir, 'ManageEngine_KB_WITH_EMBEDDED_IMAGES.xlsx')
    print(f"      Saving to: {output_file}")
    wb.save(output_file)

    print()
    print(f"      ✓ Excel file created successfully!")
    print(f"      Solutions: {len(solutions)}")
    print(f"      Images embedded: {images_embedded}")
    print()

    return output_file

def main():
    # Create session with authentication
    session = requests.Session()
    session.auth = HTTPBasicAuth(username, password)
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    })

    # Test authentication
    print("[0/4] Testing authentication...")
    try:
        response = session.get(HELPDESK_URL, verify=False, timeout=10)
        if response.status_code == 401:
            print("      ✗ Authentication failed! Check username/password.")
            input("Press Enter to exit...")
            return
        print("      ✓ Authentication successful")
        print()
    except Exception as e:
        print(f"      ✗ Connection error: {e}")
        input("Press Enter to exit...")
        return

    # Extract image references
    image_refs = extract_embedded_images_from_csv()

    if not image_refs:
        print("✗ No image references found. Check that solutions.csv is in the same directory.")
        input("Press Enter to exit...")
        return

    # Download images
    downloaded_images = save_images_locally(image_refs, session)

    if not downloaded_images:
        print("\n✗ No images were downloaded. Check ManageEngine URL and credentials.")
        input("Press Enter to exit...")
        return

    # Create Excel with embedded images
    output_file = create_excel_with_images(downloaded_images, image_refs)

    print("[4/4] Complete!")
    print("=" * 70)
    print(f"✓ SUCCESS! Your knowledge base with embedded images is ready:")
    print(f"  {output_file}")
    print("=" * 70)
    input("\nPress Enter to exit...")

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nCancelled by user.")
        input("Press Enter to exit...")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        input("\nPress Enter to exit...")
