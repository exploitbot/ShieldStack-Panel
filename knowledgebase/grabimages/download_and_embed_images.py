#!/usr/bin/env python3
"""
Download images from ManageEngine web interface and embed them into Excel
"""
import csv
import re
import os
import sys
import time
import requests
from requests.auth import HTTPBasicAuth
import urllib3
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from io import BytesIO
from PIL import Image

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Increase CSV field size limit
csv.field_size_limit(sys.maxsize)

# ManageEngine credentials and URL
HELPDESK_URL = "https://helpdesk.appsforte.com"
# You'll need to provide credentials
print("ManageEngine ServiceDesk Image Downloader")
print("=" * 60)
username = input("Enter ManageEngine username: ").strip()
password = input("Enter ManageEngine password: ").strip()
print()

def extract_embedded_images_from_csv():
    """Extract all embedded image references from solutions CSV"""
    print("Extracting embedded image references...")

    image_refs = {}  # solution_id -> list of image_ids

    with open('/tmp/solutions.csv', 'r', encoding='utf-8') as f:
        content = f.read()

    # Find all solutions with their IDs and embedded images
    lines = content.split('\n')
    for line in lines:
        if not line.strip():
            continue

        # Extract solution ID (first field)
        match = re.match(r'^(\d+),', line)
        if not match:
            continue

        sol_id = match.group(1)

        # Find all image references in this line
        img_matches = re.findall(r'/api/v3/solutions/\d+/images/(\d+)', line)

        if img_matches:
            image_refs[sol_id] = list(set(img_matches))  # Remove duplicates

    return image_refs

def download_image(session, solution_id, image_id):
    """Download a single image from ManageEngine"""
    url = f"{HELPDESK_URL}/api/v3/solutions/{solution_id}/images/{image_id}"

    try:
        response = session.get(url, verify=False, timeout=30)

        if response.status_code == 200 and len(response.content) > 100:
            return response.content
        else:
            # Try alternate URL format
            url2 = f"{HELPDESK_URL}/solutions/{solution_id}/images/{image_id}"
            response = session.get(url2, verify=False, timeout=30)

            if response.status_code == 200 and len(response.content) > 100:
                return response.content
    except Exception as e:
        print(f"    Error downloading: {e}")

    return None

def save_images_locally(image_refs, session):
    """Download all images and save them locally"""
    image_dir = "/home/appsforte/solution_images_downloaded"
    os.makedirs(image_dir, exist_ok=True)

    print(f"\nDownloading images from {HELPDESK_URL}...")
    print("This may take several minutes...\n")

    downloaded_images = {}  # (sol_id, img_id) -> local_path
    total_images = sum(len(imgs) for imgs in image_refs.values())
    current = 0
    successful = 0

    for sol_id, img_ids in sorted(image_refs.items()):
        for img_id in img_ids:
            current += 1
            print(f"[{current}/{total_images}] Solution {sol_id}, Image {img_id}...", end=' ')

            img_data = download_image(session, sol_id, img_id)

            if img_data:
                # Save to file
                filename = f"solution_{sol_id}_image_{img_id}.png"
                filepath = os.path.join(image_dir, filename)

                try:
                    with open(filepath, 'wb') as f:
                        f.write(img_data)

                    downloaded_images[(sol_id, img_id)] = filepath
                    print("✓")
                    successful += 1
                except Exception as e:
                    print(f"✗ (save error: {e})")
            else:
                print("✗ (download failed)")

            # Be nice to the server
            time.sleep(0.5)

    print(f"\n{'='*60}")
    print(f"Downloaded {successful}/{total_images} images")
    print(f"Images saved to: {image_dir}")
    print(f"{'='*60}\n")

    return downloaded_images

def create_excel_with_images(downloaded_images, image_refs):
    """Create Excel file with embedded images"""
    print("Creating Excel spreadsheet with embedded images...")

    # Load solutions
    solutions = {}
    with open('/tmp/solutions.csv', 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            solutions[row['solutionid']] = row

    # Load attachments
    attachments = defaultdict(list)
    try:
        with open('/tmp/solution_attachments.csv', 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                attachments[row['solutionid']].append({
                    'filename': row['attachmentname'],
                })
    except:
        pass

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Knowledge Base with Images"

    # Headers
    headers = ['Solution ID', 'Title', 'Description', 'Images', 'Attachments']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30

    # Add data rows
    row_num = 2
    for sol_id in sorted(solutions.keys(), key=lambda x: int(x)):
        sol = solutions[sol_id]

        # Clean description (remove HTML tags)
        desc_text = re.sub(r'<[^>]+>', '', sol['description'])
        desc_text = re.sub(r'\s+', ' ', desc_text).strip()[:500]  # Limit length

        # Basic info
        ws.cell(row=row_num, column=1, value=int(sol_id))
        ws.cell(row=row_num, column=2, value=sol['title'])
        ws.cell(row=row_num, column=3, value=desc_text)

        # Add images in column D
        if sol_id in image_refs:
            img_ids = image_refs[sol_id]
            ws.cell(row=row_num, column=4, value=f"{len(img_ids)} images")

            # Try to embed first image
            first_img_id = img_ids[0]
            img_key = (sol_id, first_img_id)

            if img_key in downloaded_images:
                try:
                    img_path = downloaded_images[img_key]

                    # Resize image if needed
                    pil_img = Image.open(img_path)
                    max_width = 200
                    max_height = 150

                    if pil_img.width > max_width or pil_img.height > max_height:
                        pil_img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                        resized_path = img_path.replace('.png', '_resized.png')
                        pil_img.save(resized_path)
                        img_path = resized_path

                    # Add to Excel
                    excel_img = ExcelImage(img_path)
                    excel_img.anchor = f'D{row_num}'
                    ws.add_image(excel_img)

                    # Adjust row height
                    ws.row_dimensions[row_num].height = 120
                except Exception as e:
                    print(f"  Warning: Could not embed image for solution {sol_id}: {e}")
        else:
            ws.cell(row=row_num, column=4, value="No images")

        # Attachments
        sol_attachments = attachments.get(sol_id, [])
        if sol_attachments:
            att_text = '\n'.join([a['filename'] for a in sol_attachments[:3]])
            if len(sol_attachments) > 3:
                att_text += f"\n... and {len(sol_attachments)-3} more"
            ws.cell(row=row_num, column=5, value=att_text)
        else:
            ws.cell(row=row_num, column=5, value="None")

        row_num += 1

        if row_num % 50 == 0:
            print(f"  Processed {row_num-1} solutions...")

    # Freeze header
    ws.freeze_panes = 'A2'

    # Save
    output_file = '/home/appsforte/ManageEngine_KB_WITH_EMBEDDED_IMAGES.xlsx'
    print(f"Saving to {output_file}...")
    wb.save(output_file)

    print(f"\n✓ Complete! File saved: {output_file}")
    print(f"  Solutions: {len(solutions)}")
    print(f"  Images embedded: {len([k for k in downloaded_images.keys() if k[0] in image_refs])}")

def main():
    # Create session with authentication
    session = requests.Session()
    session.auth = HTTPBasicAuth(username, password)
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    })

    # Test authentication
    print("Testing authentication...")
    try:
        response = session.get(HELPDESK_URL, verify=False, timeout=10)
        if response.status_code == 401:
            print("✗ Authentication failed! Check username/password.")
            return
        print("✓ Authentication successful\n")
    except Exception as e:
        print(f"✗ Connection error: {e}")
        return

    # Extract image references
    image_refs = extract_embedded_images_from_csv()
    print(f"Found {sum(len(v) for v in image_refs.values())} embedded images across {len(image_refs)} solutions\n")

    # Download images
    downloaded_images = save_images_locally(image_refs, session)

    if not downloaded_images:
        print("\n✗ No images were downloaded. Cannot create Excel with images.")
        print("  Check that the ManageEngine URL and credentials are correct.")
        return

    # Create Excel with embedded images
    create_excel_with_images(downloaded_images, image_refs)

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nCancelled by user.")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
